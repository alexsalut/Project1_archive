#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/9/25 13:53
# @Author  : Suying
# @Site    : 
# @File    : monitor_archive.py
import os.path

import pandas as pd
import xlwings as xw
import datetime
import time
from util.utils import retry_save_excel, retry_remove_excel



def archive_monitor_today(monitor_path, remote_monitor_dir, monitor_dir, today):
    local_path = rf'{monitor_dir}/monitor_{today}.xlsx'
    if not os.path.exists(local_path):
        try:
            df = get_refreshed_data(monitor_path)
            df.index = df.index + 1
            df = df.applymap(lambda x: str(x) if isinstance(x, datetime.time) else x)
            df.columns = [chr(ord('A') + i) for i in range(len(df.columns))]
            app = xw.App(visible=False, add_book=False)
            wb = xw.books.open(monitor_path)
            sheet = wb.sheets['monitor目标持仓']

            for index in df.index:
                for col in df.columns:
                    sheet.range(f'{col}{index}').value = df.loc[index, col]

            retry_save_excel(wb=wb, file_path=local_path)
            retry_save_excel(wb=wb, file_path=rf'{remote_monitor_dir}/monitor_{today}.xlsx')

            # retry_remove_excel(file_path=monitor_path)
            # retry_remove_excel(file_path=rf'{remote_monitor_dir}/monitor_{today}_formula.xlsx')
            wb.close()
            app.quit()
            print('Archive today rolling_check successfully')
        except Exception as e:
            print(e)
            print('Retry archive today rolling_check in 10 seconds')
            time.sleep(10)
            archive_monitor_today(monitor_path, remote_monitor_dir, monitor_dir, today)
    else:
        print('Today monitor already archived')


def get_refreshed_data(monitor_path):
    df = pd.read_excel(monitor_path, sheet_name=0, index_col=None, header=None)
    if (df == 'Refreshing').any().any():
        print('Monitor data is refreshing, wait for 10 seconds to retry for archiving')
        get_refreshed_data()
    else:
        print('Monitor data refreshed and ready to be archived')
        return df


if __name__ == '__main__':


    import openpyxl

    # Load the Excel file
    workbook = openpyxl.load_workbook(r'C:\Users\Yz02\Desktop\strategy_update\monitor_20231030.xlsx')



    # Iterate through the sheets and check cell values
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                cell_value = cell.value
                if isinstance(cell_value, str) and cell_value.startswith('='):
                    print(f"Found a potential formula in cell {cell.coordinate}: {cell_value}")

    # Close the Excel file
    workbook.close()

    # Close the Excel file
    workbook.close()
